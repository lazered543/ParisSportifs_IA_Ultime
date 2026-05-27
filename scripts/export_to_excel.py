from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

EXCEL = Path("systeme_prediction_ultime.xlsx")
PRED = Path("data/predictions/predictions_today.csv")
VALUE = Path("data/predictions/value_bets_today.csv")
UPCOMING = Path("data/processed/upcoming_odds.csv")
TRACKING = Path("tracking_results.csv")
ARCHIVE = Path("data/archive/finished_bets_archive.csv")

def clear_sheet(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None

def write_df(ws, df):
    for c, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c).value = col
    clear_sheet(ws, 2)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c).value = val

def get_or_create_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

def main():
    if not EXCEL.exists():
        print("Excel introuvable.")
        return
    wb = load_workbook(EXCEL)
    if PRED.exists():
        df = pd.read_csv(PRED)
        write_df(wb["Predictions"], df)
    if VALUE.exists():
        df = pd.read_csv(VALUE)
        write_df(wb["Value_Bets"], df)
    if UPCOMING.exists():
        df = pd.read_csv(UPCOMING)
        write_df(get_or_create_sheet(wb, "Matchs_A_Venir"), df)
    if TRACKING.exists():
        df = pd.read_csv(TRACKING)
        write_df(get_or_create_sheet(wb, "Tracking"), df)
    if ARCHIVE.exists():
        df = pd.read_csv(ARCHIVE)
        write_df(get_or_create_sheet(wb, "Archive"), df)
    wb.save(EXCEL)
    print("Excel mis à jour :", EXCEL)

if __name__ == "__main__":
    main()
